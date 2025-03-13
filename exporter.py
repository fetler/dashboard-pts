import csv
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter

seen = set()

COURSES = [
    "BEng (Hons) Integrated Engineering",
    "Engineering Study Abroad",
    "Enhancing Research Skills",
    "MPhil/PhD Applied Mathematics",
    "MPhil/PhD Astrophysics",
    "MPhil/PhD Computer Science",
    "MPhil/PhD Engineering",
    "MSc by Research Applied Mathematics",
    "MSc by Research Computational Physics",
    "MSc by Research Engineering",
    "MSc Computer Science by Research",
    "PG Cert in Military Electronic Mission Protection",
    "School of Computer Science Erasmus Exchange Programme"
]

class CSVFilterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Personal Tutor Tool")

        self.file_path = tk.StringVar()
        tk.Label(root, text="Selected File:").grid(row=0, column=0, sticky="w", padx=10)
        self.file_label = tk.Label(root, textvariable=self.file_path, fg="blue")
        self.file_label.grid(row=0, column=1, sticky="w")
        tk.Button(root, text="Select CSV File", command=self.select_file).grid(row=0, column=2, padx=5, pady=5)

        self.excluded_courses = {}
        tk.Label(root, text="Excluded Courses:").grid(row=1, column=0, sticky="w", padx=10)
        tk.Label(root, text="Personal Tutors:").grid(row=9, column=0, sticky="w", padx=10, pady=5)

        self.checkbox_frame = tk.Frame(root)
        self.checkbox_frame.grid(row=2, column=0, columnspan=3, sticky="w", padx=10)
        for index, course in enumerate(COURSES):
            var = tk.BooleanVar(value=True)
            self.excluded_courses[course] = var
            tk.Checkbutton(self.checkbox_frame, text=course, variable=var).grid(row=index // 2, column=index % 2, sticky="w")
        self.pt_frame = tk.Frame(root)
        self.pt_frame.grid(row=10, column=0, columnspan=3, sticky="w", padx=10)
        self.no_tutor_var = tk.BooleanVar(value=True)
        tk.Checkbutton(self.pt_frame, text="Only show students with no personal tutor", variable=self.no_tutor_var).grid(row=3, column=0, columnspan=3, sticky="w")
        tk.Button(root, text="Run Script", command=self.run_script).grid(row=15, column=0, columnspan=3, pady=10)
        tk.Button(root, text="Export to XLSX", command=self.export_to_xlsx).grid(row=17, column=0, columnspan=3, pady=10)
        tk.Button(root, text="Exit", command=root.destroy).grid(row=19, column=0, columnspan=3, pady=10)

        self.filtered_data = []

    def select_file(self):
        file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if file:
            self.file_path.set(file)

    def run_script(self):
        file = self.file_path.get()
        if not file:
            messagebox.showerror("Error", "Please select a CSV file first")
            return

        excluded_courses = {course for course, var in self.excluded_courses.items() if var.get()}
        filter_no_tutor = self.no_tutor_var.get()

        seen = set()
        try:
            with open(file, newline='', encoding="utf-8") as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    student_id = row.get('StudentID2', '').strip()
                    course_title = row.get('CourseTitle2', '').strip()
                    personal_tutor = row.get('Textbox239', '').strip()

                    if course_title in excluded_courses:
                        continue

                    if filter_no_tutor and personal_tutor:
                        continue

                    if student_id and student_id not in seen:
                        seen.add(student_id)
                        student_data = {
                            'Student ID': student_id,
                            'First Name': row.get('FirstForename2', ''),
                            'Last Name': row.get('Surname2', ''),
                            'Course Title': row.get('CourseTitle2', ''),
                            'Course Level/Year': row.get('CourseSession', ''),
                            'Personal Tutor': personal_tutor
                        }
                        self.filtered_data.append(student_data)
                        
            self.filtered_data.sort(key=lambda x: (x['Course Level/Year'], x['Course Title'], x['Last Name']))

            for row in self.filtered_data:
                print(row['Student ID'], row['First Name'], row['Last Name'], row['Course Title'], row['Course Level/Year'], row['Personal Tutor'])

            messagebox.showinfo("Success", "Script executed successfully. Check the console output.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def export_to_xlsx(self):
        if not self.filtered_data:
            messagebox.showerror("Error", "No data to export. Run the script first.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Data"

            headers = ["Student ID", "First Name", "Last Name", "Course Title", "Course Level/Year", "Personal Tutor"]
            sheet.append(headers)

            self.filtered_data.sort(key=lambda x: (x['Course Level/Year'], x['Course Title'], x['Last Name']))

            for row in self.filtered_data:
                sheet.append([row[col] for col in headers])

            for col_idx, header in enumerate(headers, start=1):
                max_length = len(header)
                col_letter = get_column_letter(col_idx)

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                sheet.column_dimensions[col_letter].width = max_length + 2

            workbook.save(file_path)
            messagebox.showinfo("Success", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while exporting: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = CSVFilterApp(root)
    root.mainloop()
